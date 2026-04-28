import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_report():
    input_file = 'data/swiggy_delivery_data.csv'
    output_file = 'data/Swiggy_Business_Report.xlsx'

    if not os.path.exists(input_file):
        print("Error: Data file not found. Please run generate_data.py first.")
        return

    df = pd.read_csv(input_file)
    df['Order_Date'] = pd.to_datetime(df['Order_Date'])

    print("Generating Excel Report...")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:

        # --- Sheet 1: Raw Data (like screenshot) ---
        raw_cols = ['Restaurant_Name', 'Area', 'City', 'Order_Value', 'Ratings',
                    'Order_ID', 'Food_Type', 'Address', 'Delivery_Time', 'Delivery_Status']
        df[raw_cols].to_excel(writer, sheet_name='Raw_Data', index=False)

        # --- Sheet 2: City-wise Performance ---
        city_stats = df.groupby('City').agg(
            Total_Orders=('Order_ID', 'count'),
            Total_Revenue=('Order_Value', 'sum'),
            Avg_Rating=('Ratings', 'mean'),
            Avg_Delivery_Time=('Delivery_Time', 'mean')
        ).round(2).sort_values('Total_Revenue', ascending=False)
        city_stats.to_excel(writer, sheet_name='City_Performance')

        # --- Sheet 3: Restaurant Rankings ---
        rest_stats = df.groupby('Restaurant_Name').agg(
            Total_Orders=('Order_ID', 'count'),
            Total_Revenue=('Order_Value', 'sum'),
            Avg_Rating=('Ratings', 'mean'),
            Avg_Price=('Order_Value', 'mean')
        ).round(2).sort_values('Total_Revenue', ascending=False)
        rest_stats.to_excel(writer, sheet_name='Top_Restaurants')

        # --- Sheet 4: Food Type Analysis ---
        food_stats = df.groupby('Food_Type').agg(
            Total_Orders=('Order_ID', 'count'),
            Total_Revenue=('Order_Value', 'sum'),
            Avg_Rating=('Ratings', 'mean')
        ).round(2).sort_values('Total_Orders', ascending=False)
        food_stats.to_excel(writer, sheet_name='Food_Type_Analysis')

        # --- Sheet 5: Area Performance ---
        area_stats = df.groupby(['City', 'Area']).agg(
            Total_Orders=('Order_ID', 'count'),
            Avg_Price=('Order_Value', 'mean'),
            Avg_Rating=('Ratings', 'mean')
        ).round(2).sort_values('Avg_Price', ascending=False)
        area_stats.to_excel(writer, sheet_name='Area_Performance')

        # --- Sheet 6: Payment Method ---
        payment_stats = df.groupby('Payment_Method').agg(
            Total_Orders=('Order_ID', 'count'),
            Total_Revenue=('Order_Value', 'sum')
        ).round(2)
        payment_stats.to_excel(writer, sheet_name='Payment_Insights')

        # --- Sheet 7: Summary KPIs ---
        summary = pd.DataFrame({
            'Metric': ['Total Orders', 'Total Revenue (₹)', 'Avg Rating', 'Avg Price (₹)', 'Avg Delivery Time (mins)', 'Total Restaurants', 'Total Cities'],
            'Value': [
                len(df),
                round(df['Order_Value'].sum(), 2),
                round(df['Ratings'].mean(), 2),
                round(df['Order_Value'].mean(), 2),
                round(df['Delivery_Time'].mean(), 2),
                df['Restaurant_Name'].nunique(),
                df['City'].nunique()
            ]
        })
        summary.to_excel(writer, sheet_name='KPI_Summary', index=False)

    # --- Apply Formatting ---
    wb = load_workbook(output_file)
    orange_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    white_font = Font(bold=True, color="000000")
    center = Alignment(horizontal='center')

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Style Header Row
        for cell in ws[1]:
            cell.fill = orange_fill
            cell.font = white_font
            cell.alignment = center
        # Auto-fit Column Width
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

    wb.save(output_file)
    print(f"\n✅ Professional Excel Report saved: {output_file}")
    print(f"   📊 Sheets: {', '.join(wb.sheetnames)}")

if __name__ == "__main__":
    generate_excel_report()
